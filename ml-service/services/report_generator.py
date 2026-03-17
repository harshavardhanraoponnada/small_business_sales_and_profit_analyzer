"""
Report Generation Service
Generates Excel and PDF reports with sales, profit, and expense analytics
"""

import os
import pandas as pd
from datetime import datetime, timedelta
try:
    from fpdf import FPDF
except ImportError:
    from fpdf.fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
from io import BytesIO
import base64
import traceback

from utils.data_loader import DataLoader
from config.settings import Config


class ReportGenerator:
    """Generate business reports in PDF and Excel formats"""
    
    def __init__(self, data_dir=None):
        self.data_dir = data_dir or Config.DATA_DIR
        self.reports_dir = Config.REPORTS_DIR
        
        # Create reports directory if it doesn't exist
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def calculate_metrics(self, sales_df, expenses_df=None, start_date=None, end_date=None):
        """Calculate key performance metrics"""
        metrics = {
            'total_revenue': 0.0,
            'total_expense': 0.0,
            'total_profit': 0.0,
            'profit_margin': 0.0,
            'transaction_count': 0,
            'date_range': None,
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        try:
            if not sales_df.empty:
                metrics['total_revenue'] = sales_df['total'].sum() if 'total' in sales_df.columns else 0.0
                metrics['transaction_count'] = len(sales_df)
                
                if start_date or end_date:
                    date_start = start_date if start_date else sales_df['date'].min()
                    date_end = end_date if end_date else sales_df['date'].max()
                    metrics['date_range'] = f"{date_start} to {date_end}"
                else:
                    metrics['date_range'] = f"{sales_df['date'].min()} to {sales_df['date'].max()}" if 'date' in sales_df.columns else "N/A"
            
            if expenses_df is not None and not expenses_df.empty:
                metrics['total_expense'] = expenses_df['amount'].sum() if 'amount' in expenses_df.columns else 0.0
            
            # Calculate profit
            metrics['total_profit'] = metrics['total_revenue'] - metrics['total_expense']
            
            # Calculate profit margin
            if metrics['total_revenue'] > 0:
                metrics['profit_margin'] = (metrics['total_profit'] / metrics['total_revenue']) * 100
            
            # Convert to 2 decimal places
            metrics['total_revenue'] = round(metrics['total_revenue'], 2)
            metrics['total_expense'] = round(metrics['total_expense'], 2)
            metrics['total_profit'] = round(metrics['total_profit'], 2)
            metrics['profit_margin'] = round(metrics['profit_margin'], 2)
            
        except Exception as e:
            print(f"Error calculating metrics: {e}")
        
        return metrics
    
    def create_sales_trend_chart(self, sales_df):
        """Create a sales trend chart as base64 image"""
        try:
            if sales_df.empty or 'date' not in sales_df.columns or 'total' not in sales_df.columns:
                print("Warning: Empty sales data or missing columns for chart")
                return None
            
            fig, ax = plt.subplots(figsize=(12, 6))
            
            # Convert date to datetime if needed
            if sales_df['date'].dtype != 'datetime64[ns]':
                sales_df['date'] = pd.to_datetime(sales_df['date'], errors='coerce')
            
            # Daily sales trend
            daily_sales = sales_df.groupby(sales_df['date'].dt.date)['total'].sum()
            
            if len(daily_sales) == 0:
                plt.close(fig)
                return None
            
            ax.plot(daily_sales.index, daily_sales.values, marker='o', linewidth=2, markersize=5, color='#4472C4')
            ax.set_title('Sales Trend', fontsize=14, fontweight='bold')
            ax.set_xlabel('Date', fontsize=12)
            ax.set_ylabel('Revenue ($)', fontsize=12)
            ax.grid(True, alpha=0.3)
            plt.xticks(rotation=45)
            plt.tight_layout()
            
            # Save to bytes buffer
            buffer = BytesIO()
            plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
            buffer.seek(0)
            plt.close(fig)
            
            return buffer
        except Exception as e:
            print(f"Error creating sales trend chart: {e}")
            traceback.print_exc()
            return None
    
    def create_expense_distribution_chart(self, expenses_df):
        """Create an expense distribution chart"""
        try:
            if expenses_df.empty or 'category' not in expenses_df.columns or 'amount' not in expenses_df.columns:
                print("Warning: Empty expense data or missing columns for chart")
                return None
            
            fig, ax = plt.subplots(figsize=(10, 6))
            
            # Expense by category pie chart
            category_expenses = expenses_df.groupby('category')['amount'].sum()
            
            if len(category_expenses) == 0:
                plt.close(fig)
                return None
            
            colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8', '#F7DC6F']
            ax.pie(category_expenses.values, labels=category_expenses.index, 
                   autopct='%1.1f%%', startangle=90, colors=colors[:len(category_expenses)])
            ax.set_title('Expense Distribution by Category', fontsize=14, fontweight='bold')
            plt.tight_layout()
            
            # Save to bytes buffer
            buffer = BytesIO()
            plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
            buffer.seek(0)
            plt.close(fig)
            
            return buffer
        except Exception as e:
            print(f"Error creating expense distribution chart: {e}")
            traceback.print_exc()
            return None
    
    def export_to_excel(self, report_data, report_type='summary'):
        """Export report data to Excel format"""
        try:
            filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # Summary sheet
            ws_summary = wb.create_sheet('Summary')
            self._populate_summary_sheet(ws_summary, report_data)
            
            # Data sheet
            if 'sales_data' in report_data and not report_data['sales_data'].empty:
                ws_data = wb.create_sheet('Sales Data')
                self._populate_data_sheet(ws_data, report_data['sales_data'])
            
            # Expenses sheet
            if 'expenses_data' in report_data and not report_data['expenses_data'].empty:
                ws_expenses = wb.create_sheet('Expenses')
                self._populate_data_sheet(ws_expenses, report_data['expenses_data'])
            
            wb.save(filepath)
            return filepath
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return None
    
    def _populate_summary_sheet(self, worksheet, report_data):
        """Populate summary metrics sheet"""
        try:
            metrics = report_data.get('metrics', {})
            
            # Headers
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            # Add report title and date
            ws = worksheet
            ws['A1'] = 'Report Summary'
            ws['A1'].font = Font(bold=True, size=14)
            
            ws['A2'] = 'Generated'
            ws['B2'] = metrics.get('generated_at', '')
            
            # Metrics section
            row = 4
            ws[f'A{row}'] = 'Key Metrics'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            
            row = 5
            metrics_list = [
                ('Total Revenue', metrics.get('total_revenue', 0)),
                ('Total Expense', metrics.get('total_expense', 0)),
                ('Total Profit', metrics.get('total_profit', 0)),
                ('Profit Margin (%)', metrics.get('profit_margin', 0)),
                ('Transaction Count', metrics.get('transaction_count', 0)),
                ('Date Range', metrics.get('date_range', 'N/A')),
            ]
            
            for label, value in metrics_list:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                
                # Format currency columns
                if 'Margin' not in label and 'Count' not in label and 'Range' not in label:
                    ws[f'B{row}'].number_format = '$#,##0.00'
                elif 'Margin' in label:
                    ws[f'B{row}'].number_format = '0.00%'
                
                row += 1
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
        except Exception as e:
            print(f"Error populating summary sheet: {e}")
    
    def _populate_data_sheet(self, worksheet, data_df):
        """Populate data sheet with raw data"""
        try:
            # Write headers
            for col_idx, col_name in enumerate(data_df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            
            # Write data
            for r_idx, row in enumerate(dataframe_to_rows(data_df, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=r_idx, column=c_idx)
                    cell.value = value
                    
                    # Format numeric columns
                    if isinstance(value, (int, float)) and 'amount' in data_df.columns:
                        cell.number_format = '$#,##0.00'
            
            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        except Exception as e:
            print(f"Error populating data sheet: {e}")
    
    def export_to_pdf(self, report_data, report_type='summary', include_charts=True):
        """Export report data to PDF format"""
        try:
            filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = os.path.join(self.reports_dir, filename)
            
            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)
            
            # Title
            pdf.set_font('Arial', 'B', 16)
            pdf.cell(0, 10, f'{report_type.replace("_", " ").title()} Report', ln=True, align='C')
            pdf.ln(5)
            
            # Metrics section
            metrics = report_data.get('metrics', {})
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, 'Key Metrics', ln=True)
            pdf.set_font('Arial', '', 10)
            
            metrics_list = [
                f"Total Revenue: ${metrics.get('total_revenue', 0):,.2f}",
                f"Total Expense: ${metrics.get('total_expense', 0):,.2f}",
                f"Total Profit: ${metrics.get('total_profit', 0):,.2f}",
                f"Profit Margin: {metrics.get('profit_margin', 0):.2f}%",
                f"Transaction Count: {metrics.get('transaction_count', 0)}",
                f"Date Range: {metrics.get('date_range', 'N/A')}",
                f"Generated: {metrics.get('generated_at', 'N/A')}",
            ]
            
            for metric in metrics_list:
                pdf.cell(0, 8, metric, ln=True)
            
            pdf.ln(10)
            
            # Add charts if requested
            if include_charts:
                if 'sales_chart' in report_data and report_data['sales_chart']:
                    try:
                        pdf.add_page()
                        pdf.set_font('Arial', 'B', 12)
                        pdf.cell(0, 10, 'Sales Trend', ln=True)
                        pdf.ln(5)
                        
                        chart_buffer = report_data['sales_chart']
                        chart_buffer.seek(0)
                        temp_chart_path = os.path.join(self.reports_dir, 'temp_sales_chart.png')
                        with open(temp_chart_path, 'wb') as f:
                            f.write(chart_buffer.read())
                        
                        # Add image to PDF (x, y, width)
                        pdf.image(temp_chart_path, x=10, y=None, w=190)
                        
                        # Cleanup temp file
                        if os.path.exists(temp_chart_path):
                            os.remove(temp_chart_path)
                    except Exception as e:
                        print(f"Error adding sales chart to PDF: {e}")
                        traceback.print_exc()
                        pdf.set_font('Arial', 'I', 10)
                        pdf.cell(0, 8, 'Chart could not be generated', ln=True)
                
                if 'expense_chart' in report_data and report_data['expense_chart']:
                    try:
                        pdf.add_page()
                        pdf.set_font('Arial', 'B', 12)
                        pdf.cell(0, 10, 'Expense Distribution', ln=True)
                        pdf.ln(5)
                        
                        chart_buffer = report_data['expense_chart']
                        chart_buffer.seek(0)
                        temp_chart_path = os.path.join(self.reports_dir, 'temp_expense_chart.png')
                        with open(temp_chart_path, 'wb') as f:
                            f.write(chart_buffer.read())
                        
                        # Add image to PDF
                        pdf.image(temp_chart_path, x=10, y=None, w=190)
                        
                        # Cleanup temp file
                        if os.path.exists(temp_chart_path):
                            os.remove(temp_chart_path)
                    except Exception as e:
                        print(f"Error adding expense chart to PDF: {e}")
                        traceback.print_exc()
                        pdf.set_font('Arial', 'I', 10)
                        pdf.cell(0, 8, 'Chart could not be generated', ln=True)
            
            pdf.output(filepath)
            print(f"PDF generated successfully: {filepath}")
            return filepath
        except Exception as e:
            print(f"Error exporting to PDF: {e}")
            traceback.print_exc()
            return None
    
    def generate_report(self, report_type='summary', format='pdf', start_date=None, end_date=None):
        """Generate a complete report with all data and metrics"""
        try:
            print(f"Generating {format} report of type '{report_type}'...")
            print(f"Data directory: {self.data_dir}")
            print(f"Reports directory: {self.reports_dir}")
            
            # Load data
            sales_df = DataLoader.load_raw_sales_data(self.data_dir, start_date, end_date)
            expenses_df = DataLoader.load_raw_expenses_data(self.data_dir, start_date, end_date)
            
            print(f"Loaded {len(sales_df)} sales records and {len(expenses_df)} expense records")
            
            # Calculate metrics
            metrics = self.calculate_metrics(sales_df, expenses_df, start_date, end_date)
            print(f"Calculated metrics: Revenue=${metrics.get('total_revenue', 0)}, Profit=${metrics.get('total_profit', 0)}")
            
            # Prepare report data
            report_data = {
                'metrics': metrics,
                'sales_data': sales_df,
                'expenses_data': expenses_df,
            }
            
            # Generate charts for PDF
            if format == 'pdf':
                print("Generating charts...")
                report_data['sales_chart'] = self.create_sales_trend_chart(sales_df)
                report_data['expense_chart'] = self.create_expense_distribution_chart(expenses_df)
                print(f"Charts generated: sales={report_data['sales_chart'] is not None}, expense={report_data['expense_chart'] is not None}")
            
            # Export based on format
            if format.lower() == 'xlsx':
                filepath = self.export_to_excel(report_data, report_type)
            elif format.lower() == 'pdf':
                filepath = self.export_to_pdf(report_data, report_type)
            else:
                raise ValueError(f"Unsupported format: {format}")
            
            if filepath and os.path.exists(filepath):
                print(f"Report generated successfully: {filepath}")
                return {
                    'success': True,
                    'filepath': filepath,
                    'filename': os.path.basename(filepath),
                    'size': os.path.getsize(filepath),
                    'metrics': metrics
                }
            else:
                error_msg = 'Failed to generate report file'
                print(f"Error: {error_msg}")
                return {
                    'success': False,
                    'error': error_msg
                }
        except Exception as e:
            error_msg = f"Error generating report: {e}"
            print(error_msg)
            traceback.print_exc()
            return {
                'success': False,
                'error': str(e)
            }

